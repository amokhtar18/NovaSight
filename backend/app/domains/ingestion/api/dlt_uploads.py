"""
NovaSight Ingestion Domain — File Upload API
=============================================

Endpoints for uploading flat files / spreadsheets to the tenant's S3 bucket
under the ``raw_uploads/`` prefix. The resulting object_key is later
referenced by file-source dlt pipelines.

Endpoints
---------
POST   /api/v1/dlt/uploads             multipart/form-data file upload
DELETE /api/v1/dlt/uploads/<object_key> remove an uploaded file
"""

from __future__ import annotations

import io
import logging
import os
import re
import uuid
from typing import Any, Dict, List, Optional

import boto3
from botocore.client import Config as BotoConfig
from botocore.exceptions import ClientError
from flask import Blueprint, current_app, jsonify, request

from app.platform.auth.decorators import authenticated
from app.platform.tenant.context import get_current_tenant_id
from app.errors import ValidationError, NotFoundError

logger = logging.getLogger(__name__)

dlt_uploads_bp = Blueprint("dlt_uploads", __name__, url_prefix="/dlt/uploads")


# Configurable limits / allowed types
MAX_UPLOAD_BYTES = int(os.environ.get("DLT_UPLOAD_MAX_BYTES", str(500 * 1024 * 1024)))

_EXT_TO_FORMAT: Dict[str, str] = {
    ".csv": "csv",
    ".tsv": "tsv",
    ".xlsx": "xlsx",
    ".xls": "xls",
    ".parquet": "parquet",
    ".json": "json",
    ".jsonl": "jsonl",
    ".ndjson": "jsonl",
}

_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._-]")
_OBJECT_KEY_RE = re.compile(r"^raw_uploads/[A-Za-z0-9._\-/]+$")


def _safe_filename(name: str) -> str:
    """Sanitize an uploaded filename."""
    base = os.path.basename(name or "")
    base = _SAFE_NAME_RE.sub("_", base).strip("._-") or "upload"
    return base[:200]


def _detect_format(filename: str) -> Optional[str]:
    ext = os.path.splitext(filename.lower())[1]
    return _EXT_TO_FORMAT.get(ext)


def _get_tenant_s3() -> Dict[str, Any]:
    """Return decrypted tenant S3 settings.

    Falls back to env-based MinIO/S3 defaults when no per-tenant
    ``object_storage`` infrastructure config row exists. The default
    bucket name is ``tenant-<slug>``.
    """
    tenant_id = get_current_tenant_id()
    if not tenant_id:
        raise ValidationError("Tenant context required")

    settings: Dict[str, Any] = {}
    try:
        from app.domains.tenants.infrastructure.config_service import (
            InfrastructureConfigService,
        )

        service = InfrastructureConfigService()
        settings = (
            service.get_effective_settings(
                service_type="object_storage",
                tenant_id=str(tenant_id),
            )
            or {}
        )
    except Exception as e:
        logger.debug(
            "Falling back to env defaults for object_storage (tenant=%s): %s",
            tenant_id,
            e,
        )

    bucket = settings.get("bucket")
    if not bucket:
        # Resolve tenant slug for default bucket name.
        slug: Optional[str] = None
        try:
            from app.domains.tenants.domain.models import Tenant

            tenant = Tenant.query.filter_by(id=tenant_id).first()
            if tenant:
                slug = tenant.slug
        except Exception as e:  # pragma: no cover - defensive
            logger.debug("Failed to resolve tenant slug: %s", e)
        slug = slug or str(tenant_id)
        bucket = f"tenant-{slug}"

    return {
        "bucket": bucket,
        "endpoint_url": settings.get("endpoint_url")
        or current_app.config.get("MINIO_ENDPOINT_URL")
        or None,
        "region": settings.get("region")
        or current_app.config.get("OBJECT_STORAGE_DEFAULT_REGION")
        or "us-east-1",
        "access_key": settings.get("access_key")
        or current_app.config.get("MINIO_ROOT_USER", ""),
        "secret_key": settings.get("secret_key")
        or current_app.config.get("MINIO_ROOT_PASSWORD", ""),
    }


def _ensure_bucket(client, bucket: str) -> None:
    """Create the bucket if it does not exist (idempotent)."""
    try:
        client.head_bucket(Bucket=bucket)
        return
    except ClientError as e:
        code = (e.response or {}).get("Error", {}).get("Code", "")
        if code not in ("404", "NoSuchBucket", "NotFound"):
            # 403 or other — surface as upload failure
            logger.warning("head_bucket failed for %s: %s", bucket, e)
            raise
    try:
        client.create_bucket(Bucket=bucket)
        logger.info("Created object storage bucket: %s", bucket)
    except ClientError as e:
        code = (e.response or {}).get("Error", {}).get("Code", "")
        if code in ("BucketAlreadyOwnedByYou", "BucketAlreadyExists"):
            return
        raise


def _build_s3_client(s3: Dict[str, Any]):
    return boto3.client(
        "s3",
        endpoint_url=s3.get("endpoint_url"),
        aws_access_key_id=s3["access_key"],
        aws_secret_access_key=s3["secret_key"],
        region_name=s3["region"],
        config=BotoConfig(signature_version="s3v4"),
    )


def _excel_sheets(buf: bytes) -> List[str]:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return list(names)
    except Exception as e:  # pragma: no cover - defensive
        logger.debug("Failed to enumerate Excel sheets: %s", e)
        return []


# Maximum number of preview rows to return to the wizard.
PREVIEW_ROWS = 20


def _df_to_preview(df) -> Dict[str, Any]:
    """Convert a small DataFrame into JSON-safe columns + rows."""
    try:
        df = df.head(PREVIEW_ROWS)
        # Replace NaN/NaT with None so JSON serialization is clean.
        safe = df.where(df.notna(), None)
        rows = safe.to_dict(orient="records")
        # Stringify any non-JSON-native scalars (datetimes, Decimals, bytes).
        cleaned: List[Dict[str, Any]] = []
        for r in rows:
            out: Dict[str, Any] = {}
            for k, v in r.items():
                if v is None or isinstance(v, (str, int, float, bool)):
                    out[str(k)] = v
                else:
                    try:
                        out[str(k)] = str(v)
                    except Exception:
                        out[str(k)] = None
            cleaned.append(out)
        return {
            "columns": [str(c) for c in df.columns.tolist()],
            "rows": cleaned,
        }
    except Exception as e:
        logger.debug("Failed to build preview from dataframe: %s", e)
        return {"columns": [], "rows": []}


def _preview_csv(buf: bytes, delimiter: str = ",") -> Dict[str, Any]:
    try:
        import pandas as pd

        df = pd.read_csv(io.BytesIO(buf), sep=delimiter, nrows=PREVIEW_ROWS)
        return _df_to_preview(df)
    except Exception as e:
        logger.debug("CSV/TSV preview failed: %s", e)
        return {"columns": [], "rows": []}


def _preview_excel(buf: bytes, sheet: Optional[str] = None) -> Dict[str, Any]:
    try:
        import pandas as pd

        df = pd.read_excel(
            io.BytesIO(buf),
            sheet_name=sheet if sheet else 0,
            nrows=PREVIEW_ROWS,
            engine="openpyxl",
        )
        return _df_to_preview(df)
    except Exception as e:
        logger.debug("Excel preview failed: %s", e)
        return {"columns": [], "rows": []}


def _preview_parquet(buf: bytes) -> Dict[str, Any]:
    try:
        import pyarrow.parquet as pq

        table = pq.read_table(io.BytesIO(buf))
        df = table.slice(0, PREVIEW_ROWS).to_pandas()
        return _df_to_preview(df)
    except Exception as e:
        logger.debug("Parquet preview failed: %s", e)
        return {"columns": [], "rows": []}


def _preview_json(buf: bytes) -> Dict[str, Any]:
    try:
        import pandas as pd

        # Try line-delimited first; fall back to a JSON array of objects.
        text = buf.decode("utf-8", errors="replace")
        try:
            df = pd.read_json(io.StringIO(text), lines=True, nrows=PREVIEW_ROWS)
        except Exception:
            df = pd.read_json(io.StringIO(text))
            df = df.head(PREVIEW_ROWS)
        return _df_to_preview(df)
    except Exception as e:
        logger.debug("JSON preview failed: %s", e)
        return {"columns": [], "rows": []}


def _preview_jsonl(buf: bytes) -> Dict[str, Any]:
    try:
        import pandas as pd

        text = buf.decode("utf-8", errors="replace")
        df = pd.read_json(io.StringIO(text), lines=True, nrows=PREVIEW_ROWS)
        return _df_to_preview(df)
    except Exception as e:
        logger.debug("JSONL preview failed: %s", e)
        return {"columns": [], "rows": []}


def _build_preview(buf: bytes, file_format: str) -> Dict[str, Any]:
    """Return ``{columns, rows}`` for the given file format and bytes buffer."""
    if file_format in ("csv", "tsv"):
        return _preview_csv(buf, delimiter="\t" if file_format == "tsv" else ",")
    if file_format in ("xlsx", "xls"):
        return _preview_excel(buf)
    if file_format == "parquet":
        return _preview_parquet(buf)
    if file_format == "json":
        return _preview_json(buf)
    if file_format == "jsonl":
        return _preview_jsonl(buf)
    return {"columns": [], "rows": []}


@dlt_uploads_bp.route("", methods=["POST"])
@authenticated
def upload_file():
    """Upload a file to ``s3://<tenant-bucket>/raw_uploads/<uuid>/<safe_name>``."""
    tenant_id = get_current_tenant_id()
    if not tenant_id:
        raise ValidationError("Tenant context required")

    if "file" not in request.files:
        raise ValidationError("No file provided (expected multipart field 'file')")

    upload = request.files["file"]
    if not upload or not upload.filename:
        raise ValidationError("Empty upload")

    original_filename = upload.filename
    file_format = _detect_format(original_filename)
    if not file_format:
        raise ValidationError(
            "Unsupported file type. Allowed: csv, tsv, xlsx, xls, parquet, json, jsonl"
        )

    safe_name = _safe_filename(original_filename)
    object_key = f"raw_uploads/{uuid.uuid4()}/{safe_name}"

    # Determine size and reject oversized uploads.
    upload.stream.seek(0, os.SEEK_END)
    size = upload.stream.tell()
    upload.stream.seek(0)
    if size <= 0:
        raise ValidationError("Empty file")
    if size > MAX_UPLOAD_BYTES:
        raise ValidationError(
            f"File too large ({size} bytes; limit is {MAX_UPLOAD_BYTES})"
        )

    # Read a peek buffer for preview (always the head; for small files this is
    # the entire payload so Excel/Parquet metadata footers are reachable).
    PEEK_LIMIT = 16 * 1024 * 1024  # 16 MiB
    peek = upload.stream.read(min(size, PEEK_LIMIT))
    upload.stream.seek(0)

    s3 = _get_tenant_s3()
    client = _build_s3_client(s3)

    # Ensure the destination bucket exists (auto-provision in dev / MinIO).
    try:
        _ensure_bucket(client, s3["bucket"])
    except Exception as e:
        logger.exception(
            "Failed to ensure bucket exists for tenant=%s bucket=%s",
            tenant_id,
            s3["bucket"],
        )
        raise ValidationError(f"Object storage bucket unavailable: {e}")

    # Stream to S3 with multipart upload.
    try:
        client.upload_fileobj(
            Fileobj=upload.stream,
            Bucket=s3["bucket"],
            Key=object_key,
            ExtraArgs={"ContentType": upload.mimetype or "application/octet-stream"},
        )
    except Exception as e:
        logger.exception("S3 upload failed for tenant=%s key=%s", tenant_id, object_key)
        raise ValidationError(f"S3 upload failed: {e}")

    # Best-effort metadata + row preview from the in-memory peek buffer.
    sheets: List[str] = []
    columns_preview: List[str] = []
    rows_preview: List[Dict[str, Any]] = []
    try:
        if file_format in {"xlsx", "xls"}:
            sheets = _excel_sheets(peek)
        preview = _build_preview(peek, file_format)
        columns_preview = preview.get("columns", [])
        rows_preview = preview.get("rows", [])
    except Exception as e:  # pragma: no cover - non-fatal
        logger.debug("Metadata peek failed: %s", e)

    return (
        jsonify(
            {
                "object_key": object_key,
                "bucket": s3["bucket"],
                "size_bytes": size,
                "file_format": file_format,
                "original_filename": original_filename,
                "sheets": sheets,
                "columns_preview": columns_preview,
                "rows_preview": rows_preview,
            }
        ),
        201,
    )


@dlt_uploads_bp.route("/<path:object_key>", methods=["DELETE"])
@authenticated
def delete_file(object_key: str):
    """Delete a previously uploaded file (must live under ``raw_uploads/``)."""
    if not _OBJECT_KEY_RE.match(object_key):
        raise ValidationError("object_key must start with 'raw_uploads/'")

    s3 = _get_tenant_s3()
    client = _build_s3_client(s3)

    try:
        client.delete_object(Bucket=s3["bucket"], Key=object_key)
    except Exception as e:
        logger.exception("S3 delete failed for key=%s", object_key)
        raise ValidationError(f"S3 delete failed: {e}")

    return jsonify({"deleted": True, "object_key": object_key}), 200
