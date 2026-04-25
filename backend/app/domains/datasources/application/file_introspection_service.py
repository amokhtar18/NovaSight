"""
NovaSight Data Sources — File Introspection Service
=====================================================

Inspects uploaded files to extract schema metadata:
- CSV/TSV: delimiter, encoding, header row, column types
- Excel: sheet names, header row per sheet, column types
- SQLite: tables, columns, row counts
- JSON: top-level structure, field types
- Parquet: schema from metadata

Canonical location: ``app.domains.datasources.application.file_introspection_service``
"""

import csv
import io
import json
import logging
import os
import sqlite3
import tempfile
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# Maximum rows to preview
PREVIEW_ROWS = 20


class FileIntrospectionService:
    """Extracts metadata and preview data from uploaded files."""

    def introspect(self, file_bytes: bytes, detected_format: str, extension: str) -> Dict[str, Any]:
        """
        Introspect a file and return metadata + preview.

        Args:
            file_bytes: Raw file content
            detected_format: Format string from validation ("csv", "excel", etc.)
            extension: Original file extension (e.g., ".csv")

        Returns:
            Dict with keys: format, metadata, columns, preview_rows
        """
        handler = {
            "csv": self._introspect_csv,
            "tsv": self._introspect_csv,
            "json": self._introspect_json,
            "parquet": self._introspect_parquet,
            "excel": self._introspect_excel,
            "sqlite": self._introspect_sqlite,
        }.get(detected_format)

        if handler is None:
            return {
                "format": detected_format,
                "metadata": {},
                "columns": [],
                "preview_rows": [],
            }

        return handler(file_bytes, extension)

    # ── CSV / TSV ─────────────────────────────────────────────────

    def _introspect_csv(self, file_bytes: bytes, extension: str) -> Dict[str, Any]:
        # Detect encoding
        encoding = self._detect_encoding(file_bytes)
        text = file_bytes.decode(encoding, errors="replace")

        # Sniff delimiter
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "\t" if extension in (".tsv",) else ","

        # Detect header
        try:
            has_header = csv.Sniffer().has_header(sample)
        except csv.Error:
            has_header = True

        # Read rows
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        all_rows = []
        for i, row in enumerate(reader):
            all_rows.append(row)
            if i >= PREVIEW_ROWS + 1:
                break

        if not all_rows:
            return {
                "format": "csv",
                "metadata": {"delimiter": delimiter, "encoding": encoding, "has_header": has_header},
                "columns": [],
                "preview_rows": [],
            }

        if has_header:
            columns = [{"name": col.strip() or f"column_{i}", "index": i} for i, col in enumerate(all_rows[0])]
            preview = all_rows[1: PREVIEW_ROWS + 1]
        else:
            columns = [{"name": f"column_{i}", "index": i} for i in range(len(all_rows[0]))]
            preview = all_rows[:PREVIEW_ROWS]

        # Infer types from preview
        for col in columns:
            col["inferred_type"] = self._infer_column_type(
                [row[col["index"]] for row in preview if col["index"] < len(row)]
            )

        return {
            "format": "csv",
            "metadata": {
                "delimiter": delimiter,
                "encoding": encoding,
                "has_header": has_header,
                "total_columns": len(columns),
            },
            "columns": columns,
            "preview_rows": preview[:10],
        }

    # ── JSON ──────────────────────────────────────────────────────

    def _introspect_json(self, file_bytes: bytes, extension: str) -> Dict[str, Any]:
        data = json.loads(file_bytes)

        if isinstance(data, list):
            records = data[:PREVIEW_ROWS]
            structure = "array_of_objects" if records and isinstance(records[0], dict) else "array"
        elif isinstance(data, dict):
            # Check if it's a single record or has a data key
            for key in ("data", "records", "rows", "items", "results"):
                if key in data and isinstance(data[key], list):
                    records = data[key][:PREVIEW_ROWS]
                    structure = f"object_with_{key}"
                    break
            else:
                records = [data]
                structure = "single_object"
        else:
            return {
                "format": "json",
                "metadata": {"structure": "scalar"},
                "columns": [],
                "preview_rows": [],
            }

        # Extract columns from first records
        if records and isinstance(records[0], dict):
            all_keys = set()
            for r in records:
                if isinstance(r, dict):
                    all_keys.update(r.keys())
            columns = [
                {"name": k, "inferred_type": self._infer_json_type(records, k)}
                for k in sorted(all_keys)
            ]
        else:
            columns = []

        preview = records[:10] if records else []

        return {
            "format": "json",
            "metadata": {"structure": structure, "total_columns": len(columns)},
            "columns": columns,
            "preview_rows": preview,
        }

    # ── Parquet ───────────────────────────────────────────────────

    def _introspect_parquet(self, file_bytes: bytes, extension: str) -> Dict[str, Any]:
        import pyarrow.parquet as pq

        pf = pq.ParquetFile(io.BytesIO(file_bytes))
        schema = pf.schema_arrow
        metadata = pf.metadata

        columns = []
        for i in range(len(schema)):
            field = schema.field(i)
            columns.append({
                "name": field.name,
                "data_type": str(field.type),
                "nullable": field.nullable,
            })

        # Read preview
        table = pf.read_row_groups([0]) if metadata.num_row_groups > 0 else None
        preview = []
        if table is not None:
            for batch in table.to_batches():
                for row in batch.to_pydict().values():
                    break
                df_preview = table.slice(0, min(10, table.num_rows)).to_pydict()
                # Convert to list of rows
                keys = list(df_preview.keys())
                n_rows = len(df_preview[keys[0]]) if keys else 0
                preview = [
                    {k: df_preview[k][i] for k in keys}
                    for i in range(n_rows)
                ]
                break

        return {
            "format": "parquet",
            "metadata": {
                "num_rows": metadata.num_rows,
                "num_row_groups": metadata.num_row_groups,
                "total_columns": len(columns),
            },
            "columns": columns,
            "preview_rows": preview,
        }

    # ── Excel ─────────────────────────────────────────────────────

    def _introspect_excel(self, file_bytes: bytes, extension: str) -> Dict[str, Any]:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(list(row))
                if i >= PREVIEW_ROWS + 1:
                    break

            if not rows:
                sheets.append({
                    "name": sheet_name,
                    "columns": [],
                    "preview_rows": [],
                    "has_header": True,
                    "header_row": 1,
                })
                continue

            # Assume first row is header
            columns = [
                {
                    "name": str(col).strip() if col else f"column_{j}",
                    "index": j,
                    "inferred_type": self._infer_column_type(
                        [str(r[j]) for r in rows[1:PREVIEW_ROWS + 1] if j < len(r) and r[j] is not None]
                    ),
                }
                for j, col in enumerate(rows[0])
            ]

            preview = [
                [self._serialize_cell(c) for c in row]
                for row in rows[1:11]
            ]

            sheets.append({
                "name": sheet_name,
                "columns": columns,
                "preview_rows": preview,
                "has_header": True,
                "header_row": 1,
            })

        wb.close()

        return {
            "format": "excel",
            "metadata": {
                "sheet_count": len(sheets),
                "sheet_names": [s["name"] for s in sheets],
            },
            "sheets": sheets,
            "columns": sheets[0]["columns"] if sheets else [],
            "preview_rows": sheets[0]["preview_rows"] if sheets else [],
        }

    # ── SQLite ────────────────────────────────────────────────────

    def _introspect_sqlite(self, file_bytes: bytes, extension: str) -> Dict[str, Any]:
        fd, tmp_path = tempfile.mkstemp(suffix=".sqlite")
        try:
            os.write(fd, file_bytes)
            os.close(fd)

            conn = sqlite3.connect(f"file:{tmp_path}?mode=ro", uri=True)
            try:
                cursor = conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
                )
                table_names = [row[0] for row in cursor.fetchall()]

                tables = []
                for tname in table_names:
                    # Get columns
                    col_cursor = conn.execute(f'PRAGMA table_info("{tname}")')
                    columns = []
                    for col_row in col_cursor.fetchall():
                        columns.append({
                            "name": col_row[1],
                            "data_type": col_row[2],
                            "nullable": not col_row[3],
                            "primary_key": bool(col_row[5]),
                            "default_value": col_row[4],
                        })

                    # Row count
                    count_cursor = conn.execute(f'SELECT COUNT(*) FROM "{tname}"')
                    row_count = count_cursor.fetchone()[0]

                    # Preview
                    preview_cursor = conn.execute(f'SELECT * FROM "{tname}" LIMIT 10')
                    col_names = [desc[0] for desc in preview_cursor.description]
                    preview_rows = [
                        dict(zip(col_names, row))
                        for row in preview_cursor.fetchall()
                    ]

                    tables.append({
                        "name": tname,
                        "columns": columns,
                        "row_count": row_count,
                        "preview_rows": preview_rows,
                    })

                return {
                    "format": "sqlite",
                    "metadata": {
                        "table_count": len(tables),
                        "table_names": table_names,
                    },
                    "tables": tables,
                    "columns": tables[0]["columns"] if tables else [],
                    "preview_rows": tables[0]["preview_rows"] if tables else [],
                }
            finally:
                conn.close()
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    # ── Helpers ────────────────────────────────────────────────────

    def _detect_encoding(self, file_bytes: bytes) -> str:
        """Detect text encoding."""
        # Check BOM
        if file_bytes[:3] == b"\xef\xbb\xbf":
            return "utf-8-sig"
        if file_bytes[:2] in (b"\xff\xfe", b"\xfe\xff"):
            return "utf-16"

        try:
            import chardet
            result = chardet.detect(file_bytes[:10000])
            if result["confidence"] > 0.7:
                return result["encoding"]
        except ImportError:
            pass

        # Default fallback
        try:
            file_bytes[:4096].decode("utf-8")
            return "utf-8"
        except UnicodeDecodeError:
            return "latin-1"

    def _infer_column_type(self, values: List[str]) -> str:
        """Infer column type from sample string values."""
        if not values:
            return "string"

        int_count = float_count = bool_count = 0
        for v in values:
            v = v.strip()
            if not v:
                continue
            if v.lower() in ("true", "false"):
                bool_count += 1
                continue
            try:
                int(v)
                int_count += 1
                continue
            except ValueError:
                pass
            try:
                float(v)
                float_count += 1
                continue
            except ValueError:
                pass

        total = len([v for v in values if v.strip()])
        if total == 0:
            return "string"
        if int_count == total:
            return "integer"
        if (int_count + float_count) == total:
            return "float"
        if bool_count == total:
            return "boolean"
        return "string"

    def _infer_json_type(self, records: list, key: str) -> str:
        """Infer type of a JSON field from sample records."""
        types = set()
        for r in records:
            if isinstance(r, dict) and key in r:
                v = r[key]
                if v is None:
                    continue
                if isinstance(v, bool):
                    types.add("boolean")
                elif isinstance(v, int):
                    types.add("integer")
                elif isinstance(v, float):
                    types.add("float")
                elif isinstance(v, (list, dict)):
                    types.add("object")
                else:
                    types.add("string")

        if not types:
            return "string"
        if types == {"integer"}:
            return "integer"
        if types <= {"integer", "float"}:
            return "float"
        if types == {"boolean"}:
            return "boolean"
        return "string"

    def _serialize_cell(self, value: Any) -> Any:
        """Safely serialize an Excel cell value for JSON."""
        if value is None:
            return None
        from datetime import datetime, date, time
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.isoformat()
        return value
