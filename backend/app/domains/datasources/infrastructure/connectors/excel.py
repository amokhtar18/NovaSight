"""
NovaSight Data Sources — Excel Connector
=========================================

Connector for Excel workbooks (.xlsx, .xls) stored in tenant-scoped
local filesystem storage.

Schema model:
    - Schema:  the workbook filename (one schema per workbook)
    - Tables:  one table per worksheet

Canonical location: ``app.domains.datasources.infrastructure.connectors.excel``
"""

import io
import logging
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional

from app.domains.datasources.infrastructure.connectors.base import (
    BaseConnector,
    ConnectionConfig,
    ConnectorException,
    ConnectionTestException,
)
from app.domains.datasources.domain.value_objects import ColumnInfo, TableInfo

logger = logging.getLogger(__name__)


class ExcelConnector(BaseConnector):
    """
    Connector for Excel workbook data sources.

    Each worksheet is exposed as a table. The connector reads files using
    openpyxl in read-only, data-only mode (no formula evaluation).
    """

    connector_type = "excel"
    supported_auth_methods: List[str] = []
    supports_ssl: bool = False
    default_port: int = 0

    def __init__(self, config: ConnectionConfig):
        super().__init__(config)
        self._file_path: Optional[Path] = None
        self._wb = None  # openpyxl workbook
        self._workbook_name: str = "workbook"

    # ── Lifecycle ─────────────────────────────────────────────────

    def connect(self) -> None:
        """Locate and open the Excel workbook."""
        try:
            import openpyxl
        except ImportError:
            raise ConnectorException("openpyxl is required for Excel support. Install it: pip install openpyxl")

        file_ref = self.config.extra_params.get("file_ref")
        if not file_ref:
            raise ConnectorException("Missing 'file_ref' in extra_params")

        from app.platform.infrastructure.file_storage import FileStorageService

        parts = file_ref.split("/")
        if len(parts) < 4 or parts[0] != "tenants":
            raise ConnectorException(f"Invalid file_ref format: {file_ref}")
        tenant_id = parts[1]

        storage = FileStorageService(tenant_id)
        expected_hash = self.config.extra_params.get("file_hash")
        if expected_hash and not storage.verify_hash(file_ref, expected_hash):
            raise ConnectorException("File integrity check failed: hash mismatch")

        abs_path = storage.get_file_path(file_ref)
        if abs_path is None:
            raise ConnectorException(f"File not found: {file_ref}")

        self._file_path = abs_path
        original_name = self.config.extra_params.get("file_name", abs_path.stem)
        self._workbook_name = Path(original_name).stem or "workbook"

        self._wb = openpyxl.load_workbook(
            str(abs_path), read_only=True, data_only=True
        )
        self._is_connected = True
        logger.info(f"ExcelConnector: opened workbook {file_ref} ({len(self._wb.sheetnames)} sheets)")

    def disconnect(self) -> None:
        if self._wb is not None:
            try:
                self._wb.close()
            except Exception:
                pass
        self._wb = None
        self._file_path = None
        self._is_connected = False

    def test_connection(self) -> bool:
        try:
            if self._wb is None:
                self.connect()
            if not self._wb.sheetnames:
                raise ConnectionTestException("Excel file contains no sheets")
            return True
        except ConnectorException as e:
            raise ConnectionTestException(str(e))

    # ── Schema Introspection ───────────────────────────────────────

    def get_schemas(self) -> List[str]:
        return [self._workbook_name]

    def get_tables(self, schema: str) -> List[TableInfo]:
        if self._wb is None:
            raise ConnectorException("Not connected")
        tables = []
        for sheet_name in self._wb.sheetnames:
            columns = self._get_sheet_columns(sheet_name)
            tables.append(TableInfo(name=sheet_name, schema=schema, columns=columns))
        return tables

    def get_table_schema(self, schema: str, table: str) -> TableInfo:
        if self._wb is None:
            raise ConnectorException("Not connected")
        if table not in self._wb.sheetnames:
            raise ConnectorException(f"Sheet '{table}' not found in workbook")
        columns = self._get_sheet_columns(table)
        return TableInfo(name=table, schema=schema, columns=columns)

    # ── Data Access ────────────────────────────────────────────────

    def fetch_data(
        self,
        query: str = "",
        params: Optional[Dict[str, Any]] = None,
        batch_size: int = 10000,
    ) -> Iterator[List[Dict[str, Any]]]:
        """Stream data from the first (or specified) sheet."""
        if not self._is_connected or self._wb is None:
            raise ConnectorException("Not connected")

        # Determine which sheet to read; "query" can be a sheet name
        sheet_name = query.strip() if query.strip() in self._wb.sheetnames else self._wb.sheetnames[0]

        yield from self._read_sheet(sheet_name, batch_size)

    def validate_query(self, query: str):
        """Excel connector: query is a sheet name or empty."""
        if self._wb and query.strip() and query.strip() not in self._wb.sheetnames:
            names = ", ".join(self._wb.sheetnames)
            return False, f"Sheet '{query.strip()}' not found. Available: {names}"
        return True, ""

    # ── Private Helpers ────────────────────────────────────────────

    def _get_sheet_columns(self, sheet_name: str) -> List[ColumnInfo]:
        ws = self._wb[sheet_name]
        header_row_num = int(self.config.extra_params.get("header_row", 1))

        header_row = None
        sample_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_num = i + 1
            if row_num == header_row_num:
                header_row = row
            elif header_row is not None:
                sample_rows.append(row)
                if len(sample_rows) >= 50:
                    break

        if not header_row:
            return []

        columns = []
        for j, cell_val in enumerate(header_row):
            col_name = str(cell_val).strip() if cell_val is not None else f"column_{j}"
            if not col_name:
                col_name = f"column_{j}"
            sample_values = [
                str(r[j]) for r in sample_rows
                if j < len(r) and r[j] is not None
            ]
            data_type = self._infer_excel_type(sample_values)
            columns.append(ColumnInfo(name=col_name, data_type=data_type, nullable=True))

        return columns

    def _read_sheet(
        self, sheet_name: str, batch_size: int
    ) -> Iterator[List[Dict[str, Any]]]:
        ws = self._wb[sheet_name]
        header_row_num = int(self.config.extra_params.get("header_row", 1))

        headers: Optional[List[str]] = None
        batch: List[Dict] = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_num = i + 1
            if row_num == header_row_num:
                headers = [
                    str(c).strip() if c is not None else f"column_{j}"
                    for j, c in enumerate(row)
                ]
                continue
            if headers is None:
                continue

            record = {}
            for j, col_name in enumerate(headers):
                val = row[j] if j < len(row) else None
                record[col_name] = self._serialize_value(val)
            batch.append(record)

            if len(batch) >= batch_size:
                yield batch
                batch = []

        if batch:
            yield batch

    def _serialize_value(self, val: Any) -> Any:
        from datetime import datetime, date, time
        if val is None:
            return None
        if isinstance(val, (datetime, date, time)):
            return val.isoformat()
        return val

    def _infer_excel_type(self, values: List[str]) -> str:
        if not values:
            return "String"
        int_count = float_count = 0
        for v in values:
            try:
                int(v)
                int_count += 1
                continue
            except ValueError:
                pass
            try:
                float(v)
                float_count += 1
            except ValueError:
                pass
        total = len(values)
        if int_count == total:
            return "Int64"
        if (int_count + float_count) == total:
            return "Float64"
        return "String"
